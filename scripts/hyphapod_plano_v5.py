"""
HyphaPod — Plano v5: leyenda en lateral izquierdo, cajetín abajo derecha.
Columna izquierda: ~55mm para leyenda. Plano ocupa el resto.
"""
import re, os, sys, math
import openpyxl, ezdxf
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.units import mm
from reportlab.lib.colors import black, white, Color

_SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
_DISEÑOS_DIR   = os.path.join(_SCRIPT_DIR, "..", "diseños")
EXCEL_DISEÑO   = os.path.join(_DISEÑOS_DIR, "HyphaPod_diseño.xlsx")
EXCEL_CATALOGO = os.path.join(_SCRIPT_DIR, "catalogo_especies_v5.xlsx")
HOJA_DISEÑO    = "CUADRÍCULA HyphaPod"
HOJA_CATALOGO  = "CATÁLOGO DE ESPECIES"

CAJETIN = {
    "titulo":"PLANO DE PLANTACIÓN — HyphaPod Piloto",
    "subtitulo":"Microbosque urbano · Método Miyawaki",
    "proyecto":"MycoRoots — TFG Ingeniería Agroalimentaria",
    "autora":"Ana Velasco Márquez",
    "universidad":"Universidad de Córdoba",
    "escala":"1:20","fecha":"2026","nplano":"01",
}
COLOR_ESTRATO={"CA":"#A8D870","MI":"#4CB84C","ME":"#2A7A35","MG":"#1A4520"}
COLOR_SUELO="#DDD0A8"
A1_W=841*mm; A1_H=594*mm

def hex2c(h):
    h=h.lstrip("#"); return Color(int(h[:2],16)/255,int(h[2:4],16)/255,int(h[4:],16)/255)

def cargar_catalogo(path):
    wb=openpyxl.load_workbook(path,read_only=True); ws=wb[HOJA_CATALOGO]; cat={}
    for row in ws.iter_rows(min_row=4,values_only=True):
        cod=row[0]
        if not cod or str(cod).startswith("⚙"): continue
        cod=str(cod).strip(); estrato=str(row[3]).strip() if row[3] else cod[:2]
        cat[cod]={"estrato":estrato,"nombre_cientifico":str(row[2]).strip() if row[2] else ""}
    return cat

def extraer_codigos(t):
    return re.findall(r'\b(CA\d{2}|MI\d{2}|ME\d{2}|MG\d{2})\b',str(t)) if t else []

def leer_diseño(path,cat):
    wb=openpyxl.load_workbook(path,read_only=True); ws=wb[HOJA_DISEÑO]
    celdas={}; n_filas=0; n_cols=0
    for ri,row in enumerate(ws.iter_rows(min_row=3,values_only=True)):
        if row[0] is None: break
        try: int(str(row[0]).strip())
        except: break
        n_filas+=1
        for ci,celda in enumerate(row[1:]):
            if celda is None: continue
            cods=[c for c in extraer_codigos(celda) if c in cat]
            if cods: celdas[(ci,ri)]=cods; n_cols=max(n_cols,ci+1)
    return celdas,n_cols,n_filas

def circulo(c,cx,cy,r,cod,estrato):
    c.setFillColor(hex2c(COLOR_ESTRATO[estrato]))
    c.setStrokeColor(Color(0.08,0.15,0.08)); c.setLineWidth(0.35)
    c.circle(cx,cy,r,fill=1,stroke=1)
    fs=max(3.2,min(6.0,r*0.52))
    c.setFont("Helvetica-Bold",fs)
    c.setFillColor(white if estrato in("ME","MG") else Color(0.05,0.12,0.05))
    c.drawCentredString(cx,cy-fs*0.36,cod)

def dibujar_celda(c,ox,oy,s,cods,cat):
    n=len(cods)
    if n==1:
        r=s*0.42; circulo(c,ox+s/2,oy+s/2,r,cods[0],cat[cods[0]]["estrato"])
    elif n==2:
        c.setStrokeColor(Color(0.25,0.32,0.20,alpha=0.8)); c.setLineWidth(0.5)
        c.line(ox,oy+s,ox+s,oy); r=s*0.20
        for i,(cx,cy) in enumerate([(ox+s*0.27,oy+s*0.73),(ox+s*0.73,oy+s*0.27)]):
            circulo(c,cx,cy,r,cods[i],cat[cods[i]]["estrato"])
    else:
        cx=ox+s/2; cy=oy+s/2
        c.setStrokeColor(Color(0.25,0.32,0.20,alpha=0.8)); c.setLineWidth(0.5)
        for px,py in [(cx,oy+s),(ox,oy),(ox+s,oy)]: c.line(cx,cy,px,py)
        r=s*0.19
        for i,(pcx,pcy) in enumerate([(cx-s*0.20,cy+s*0.19),(cx+s*0.20,cy+s*0.19),(cx,cy-s*0.25)]):
            circulo(c,pcx,pcy,r,cods[i],cat[cods[i]]["estrato"])

def generar_pdf(celdas,n_cols,n_filas,cat,path_out):
    c=rl_canvas.Canvas(path_out,pagesize=(A1_W,A1_H))

    # Layout: leyenda columna izquierda fija (58mm), cajetín abajo derecha
    LEY_W   = 58*mm    # ancho columna leyenda izquierda
    MAR_L   = 8*mm     # margen exterior izquierdo
    MAR_R   = 8*mm
    MAR_T   = 10*mm
    MAR_B   = 10*mm
    CAJ_H   = 42*mm    # alto cajetín
    COTA_H  = 10*mm    # espacio para cotas

    # Área disponible para el plano
    plot_x0 = MAR_L + LEY_W + 5*mm
    plot_x1 = A1_W - MAR_R
    plot_y0 = MAR_B + CAJ_H + 5*mm
    plot_y1 = A1_H - MAR_T

    plot_w = plot_x1 - plot_x0 - COTA_H
    plot_h = plot_y1 - plot_y0 - COTA_H

    s_max = min(plot_w/n_cols, plot_h/n_filas)
    escala_real = math.ceil(1000 / (s_max/mm) / 5) * 5  # redondear hacia arriba al múltiplo de 5
    s = 1000*mm / escala_real
    parcel_w = n_cols*s; parcel_h = n_filas*s
    off_x = plot_x0 + (plot_w - parcel_w)/2
    off_y = plot_y0 + COTA_H + (plot_h - parcel_h)/2

    # Suelo
    c.setFillColor(hex2c(COLOR_SUELO))
    c.rect(off_x,off_y,parcel_w,parcel_h,fill=1,stroke=0)
    # Cuadrícula
    c.setStrokeColor(Color(0.40,0.33,0.20,alpha=0.28)); c.setLineWidth(0.2)
    for i in range(n_cols+1): c.line(off_x+i*s,off_y,off_x+i*s,off_y+parcel_h)
    for j in range(n_filas+1): c.line(off_x,off_y+j*s,off_x+parcel_w,off_y+j*s)
    # Plantas
    for (col,fila),cods in celdas.items():
        dibujar_celda(c,off_x+col*s,off_y+(n_filas-fila-1)*s,s,cods,cat)
    # Borde
    c.setStrokeColor(black); c.setLineWidth(1.2)
    c.rect(off_x,off_y,parcel_w,parcel_h,fill=0,stroke=1)
    # Cotas
    c.setFont("Helvetica",7); c.setFillColor(black)
    c.drawCentredString(off_x+parcel_w/2,off_y-7*mm,f"{n_cols} m")
    c.saveState(); c.translate(off_x-8*mm,off_y+parcel_h/2); c.rotate(90)
    c.drawCentredString(0,0,f"{n_filas} m"); c.restoreState()
    c.setFont("Helvetica",5.5); c.setFillColor(Color(0.35,0.35,0.35))
    c.drawString(off_x,off_y+parcel_h+2*mm,f"Escala 1:{escala_real}  ·  1 celda = 1×1 m²")

    # ── LEYENDA LATERAL IZQUIERDO ──────────────────────────────────────
    orden=["CA","MI","ME","MG"]
    nombres_est={"CA":"Caméfitas (<1 m)","MI":"Microfaner. (1–4 m)",
                 "ME":"Mesofaner. (3–10 m)","MG":"Megafaner. (8–30 m)"}
    conteo={}
    for cods in celdas.values():
        for cod in cods: conteo[cod]=conteo.get(cod,0)+1

    items=[]
    for est in orden:
        sub=[(cod,conteo.get(cod,0),cat[cod]) for cod in sorted(conteo) if cat[cod]["estrato"]==est]
        if sub: items.append(("H",est)); items+=sub

    # Altura disponible para leyenda: toda la altura menos márgenes
    ley_x = MAR_L
    ley_y0 = MAR_B
    ley_h_avail = A1_H - MAR_T - MAR_B
    n_rows = len(items) + 2   # +2 para título y cabecera
    rh = min(5.2*mm, (ley_h_avail - 8*mm) / n_rows)

    ley_h = n_rows * rh + 8*mm
    ley_y_start = A1_H - MAR_T - ley_h

    c.setFillColor(Color(0.97,0.97,0.95)); c.setStrokeColor(black); c.setLineWidth(0.5)
    c.rect(ley_x,ley_y_start,LEY_W,ley_h,fill=1,stroke=1)

    cy_l = ley_y_start + ley_h - 3.5*mm
    c.setFont("Helvetica-Bold",7.5); c.setFillColor(black)
    c.drawString(ley_x+2*mm,cy_l,"LEYENDA"); cy_l -= rh*0.9

    # Cabecera
    fs_h=5.5; c.setFont("Helvetica-Bold",fs_h); c.setFillColor(black)
    c.drawString(ley_x+2*mm,   cy_l,"S")        # símbolo
    c.drawString(ley_x+8*mm,   cy_l,"Cód.")
    c.drawString(ley_x+20*mm,  cy_l,"Nombre científico")
    c.drawString(ley_x+LEY_W-6*mm, cy_l,"N")
    cy_l -= 1*mm
    c.setLineWidth(0.25); c.line(ley_x+1*mm,cy_l,ley_x+LEY_W-1*mm,cy_l); cy_l -= rh*0.6

    for item in items:
        if item[0]=="H":
            est=item[1]; hc=hex2c(COLOR_ESTRATO[est])
            c.setFillColor(Color(hc.red,hc.green,hc.blue,alpha=0.25))
            c.rect(ley_x+1*mm,cy_l-1.5*mm,LEY_W-2*mm,rh,fill=1,stroke=0)
            c.setFont("Helvetica-Bold",min(6.0,rh*0.55)); c.setFillColor(black)
            c.drawString(ley_x+2*mm,cy_l,nombres_est[est]); cy_l-=rh
        else:
            cod,n,esp=item; est=esp["estrato"]
            rc=min(2.2*mm, rh*0.42)
            c.setFillColor(hex2c(COLOR_ESTRATO[est]))
            c.setStrokeColor(Color(0,0,0,alpha=0.3)); c.setLineWidth(0.2)
            c.circle(ley_x+2*mm+rc,cy_l+rh*0.28,rc,fill=1,stroke=1)
            c.setFont("Helvetica-Bold",max(3.0,rc*0.45))
            c.setFillColor(white if est in("ME","MG") else Color(0.05,0.1,0.05))
            c.drawCentredString(ley_x+2*mm+rc,cy_l+rh*0.12,cod)
            fs_t=min(5.5,rh*0.52); c.setFont("Helvetica",fs_t); c.setFillColor(black)
            c.drawString(ley_x+8*mm,cy_l,cod)
            c.setFont("Helvetica-Oblique",fs_t)
            max_chars=int((LEY_W-28*mm)/(fs_t*0.48))
            c.drawString(ley_x+20*mm,cy_l,esp["nombre_cientifico"][:max_chars])
            c.setFont("Helvetica",fs_t)
            c.drawRightString(ley_x+LEY_W-2*mm,cy_l,str(n)); cy_l-=rh

    # Barra de escala gráfica (entre leyenda de especies y subdivisión)
    _sb_h = 14*mm
    _sb_y0 = ley_y_start - 4*mm - _sb_h
    if _sb_y0 >= MAR_B:
        _bar_x = ley_x + 2*mm
        _half_max = (LEY_W - 4*mm) / 2
        if s <= _half_max:
            _seg_w, _lbl1, _lbl2 = s, "1 m", "2 m"
        else:
            _seg_w, _lbl1, _lbl2 = s/2, "0.5 m", "1 m"
        c.setFont("Helvetica-Bold",5); c.setFillColor(black)
        c.drawString(_bar_x, _sb_y0+_sb_h-2.5*mm, "ESCALA GRÁFICA")
        c.setLineWidth(0.3)
        for _i in range(2):
            c.setFillColor(black if _i%2==0 else white)
            c.setStrokeColor(black)
            c.rect(_bar_x+_i*_seg_w, _sb_y0+6*mm, _seg_w, 2.5*mm, fill=1, stroke=1)
        c.setFont("Helvetica",4); c.setFillColor(black)
        c.drawCentredString(_bar_x,           _sb_y0+4.5*mm, "0")
        c.drawCentredString(_bar_x+_seg_w,    _sb_y0+4.5*mm, _lbl1)
        c.drawCentredString(_bar_x+2*_seg_w,  _sb_y0+4.5*mm, _lbl2)
        sub_y0 = _sb_y0 - 2*mm
    else:
        sub_y0 = ley_y_start - 2*mm

    # Mini-leyenda subdivisión debajo de la leyenda de especies

    cel=9*mm; sub_h=cel*3+22*mm; sub_y_start=sub_y0-sub_h
    if sub_y_start >= MAR_B:
        c.setFillColor(Color(0.97,0.97,0.95)); c.setStrokeColor(black); c.setLineWidth(0.5)
        c.rect(ley_x,sub_y_start,LEY_W,sub_h,fill=1,stroke=1)
        c.setFont("Helvetica-Bold",6); c.setFillColor(black)
        c.drawString(ley_x+2*mm,sub_y_start+sub_h-3.5*mm,"SUBDIVISIÓN")
        for idx,(lbl,e1,e2,e3) in enumerate([("1 planta","MG",None,None),("2 plantas","MI","CA",None),("3 plantas","CA","MI","ME")]):
            yc=sub_y_start+sub_h-18*mm-idx*(cel+5*mm)
            c.setFont("Helvetica",5.5); c.setFillColor(black); c.drawString(ley_x+2*mm,yc+cel+1*mm,lbl)
            c.setStrokeColor(Color(0.2,0.3,0.2)); c.setLineWidth(0.5)
            c.rect(ley_x+2*mm,yc,cel,cel,fill=0,stroke=1)
            if e2 is None:
                c.setFillColor(hex2c(COLOR_ESTRATO[e1])); c.circle(ley_x+2*mm+cel/2,yc+cel/2,cel*0.40,fill=1,stroke=0)
            elif e3 is None:
                c.line(ley_x+2*mm,yc+cel,ley_x+2*mm+cel,yc); r2=cel*0.19
                c.setFillColor(hex2c(COLOR_ESTRATO[e1])); c.circle(ley_x+2*mm+cel*0.27,yc+cel*0.73,r2,fill=1,stroke=0)
                c.setFillColor(hex2c(COLOR_ESTRATO[e2])); c.circle(ley_x+2*mm+cel*0.73,yc+cel*0.27,r2,fill=1,stroke=0)
            else:
                ccx=ley_x+2*mm+cel/2; ccy=yc+cel/2
                for px,py in [(ccx,yc+cel),(ley_x+2*mm,yc),(ley_x+2*mm+cel,yc)]: c.line(ccx,ccy,px,py)
                r3=cel*0.18
                c.setFillColor(hex2c(COLOR_ESTRATO[e1])); c.circle(ccx-cel*0.20,ccy+cel*0.19,r3,fill=1,stroke=0)
                c.setFillColor(hex2c(COLOR_ESTRATO[e2])); c.circle(ccx+cel*0.20,ccy+cel*0.19,r3,fill=1,stroke=0)
                c.setFillColor(hex2c(COLOR_ESTRATO[e3])); c.circle(ccx,ccy-cel*0.25,r3,fill=1,stroke=0)

    # ── CAJETÍN abajo derecha ──────────────────────────────────────────
    cw=min(180*mm, A1_W-MAR_L-LEY_W-10*mm); ch=CAJ_H
    cx0=A1_W-MAR_R-cw; cy0=MAR_B
    c.setFillColor(white); c.setStrokeColor(black); c.setLineWidth(0.8)
    c.rect(cx0,cy0,cw,ch,fill=1,stroke=1)
    c.setFillColor(hex2c(COLOR_ESTRATO["MG"]))
    c.rect(cx0,cy0+ch-11*mm,cw,11*mm,fill=1,stroke=0)
    c.setFillColor(white); c.setFont("Helvetica-Bold",10)
    c.drawCentredString(cx0+cw/2,cy0+ch-7.5*mm,CAJETIN["titulo"])
    c.setFont("Helvetica",7.5)
    c.drawCentredString(cx0+cw/2,cy0+ch-11.5*mm+1*mm,CAJETIN["subtitulo"])
    col1x=cx0+cw*0.40; col2x=cx0+cw*0.70
    c.setStrokeColor(Color(0.6,0.6,0.6)); c.setLineWidth(0.3)
    c.line(col1x,cy0+1*mm,col1x,cy0+ch-11*mm)
    c.line(col2x,cy0+1*mm,col2x,cy0+ch-11*mm)
    def campo(lbl,val,x,y):
        c.setFont("Helvetica",5); c.setFillColor(Color(0.45,0.45,0.45))
        c.drawString(x+2*mm,y+3*mm,lbl.upper())
        c.setFont("Helvetica-Bold",7); c.setFillColor(black)
        c.drawString(x+2*mm,y,val)
    r1c=cy0+ch-20*mm; r2c=cy0+ch-29*mm; r3c=cy0+ch-38*mm
    campo("Proyecto",CAJETIN["proyecto"],cx0,r1c)
    campo("Universidad",CAJETIN["universidad"],col1x,r1c)
    campo("Nº Plano",CAJETIN["nplano"],col2x,r1c)
    campo("Autora",CAJETIN["autora"],cx0,r2c)
    campo("Escala",f"1:{escala_real}",col2x,r2c)
    campo("Fecha",CAJETIN["fecha"],cx0,r3c)
    total=sum(len(v) for v in celdas.values())
    campo("Total plantas",f"{total} ud.",col1x,r3c)
    campo("Formato","DIN A1 (841×594 mm)",col2x,r3c)

    c.setStrokeColor(black); c.setLineWidth(1.5)
    c.rect(5*mm,5*mm,A1_W-10*mm,A1_H-10*mm,fill=0,stroke=1)
    c.save(); print(f"  ✅  PDF: {path_out}")

def generar_dxf(celdas,n_cols,n_filas,cat,path_out):
    doc=ezdxf.new(dxfversion="R2010"); doc.units=ezdxf.units.M
    for nm,col in {"CA":3,"MI":2,"ME":4,"MG":6,"CUADRICULA":8,"BORDE":7,"COTAS":7,"DIVISION":9}.items():
        doc.layers.add(name=nm,color=col)
    msp=doc.modelspace()
    for i in range(n_cols+1): msp.add_line((i,0),(i,n_filas),dxfattribs={"layer":"CUADRICULA"})
    for j in range(n_filas+1): msp.add_line((0,j),(n_cols,j),dxfattribs={"layer":"CUADRICULA"})
    msp.add_lwpolyline([(0,0),(n_cols,0),(n_cols,n_filas),(0,n_filas),(0,0)],dxfattribs={"layer":"BORDE","lineweight":30})
    for (col,fila),cods in celdas.items():
        x0=col; y0=n_filas-fila-1; n=len(cods); cx=x0+0.5; cy=y0+0.5
        if n==1:
            cod=cods[0]; est=cat[cod]["estrato"]
            msp.add_circle((cx,cy),radius=0.40,dxfattribs={"layer":est})
            msp.add_text(cod,dxfattribs={"layer":est,"height":0.16,"halign":1,"valign":2}).set_placement((cx,cy))
        elif n==2:
            msp.add_line((x0,y0+1),(x0+1,y0),dxfattribs={"layer":"DIVISION"})
            for i,cod in enumerate(cods[:2]):
                est=cat[cod]["estrato"]; pos=(x0+0.27,y0+0.73) if i==0 else (x0+0.73,y0+0.27)
                msp.add_circle(pos,radius=0.18,dxfattribs={"layer":est})
                msp.add_text(cod,dxfattribs={"layer":est,"height":0.11,"halign":1,"valign":2}).set_placement(pos)
        else:
            for px,py in [(cx,y0+1),(x0,y0),(x0+1,y0)]: msp.add_line((cx,cy),(px,py),dxfattribs={"layer":"DIVISION"})
            for i,cod in enumerate(cods[:3]):
                est=cat[cod]["estrato"]; pos=[(cx-0.20,cy+0.19),(cx+0.20,cy+0.19),(cx,cy-0.25)][i]
                msp.add_circle(pos,radius=0.17,dxfattribs={"layer":est})
                msp.add_text(cod,dxfattribs={"layer":est,"height":0.10,"halign":1,"valign":2}).set_placement(pos)
    msp.add_linear_dim(base=(0,-1.5),p1=(0,0),p2=(n_cols,0),dxfattribs={"layer":"COTAS"}).render()
    msp.add_linear_dim(base=(-1.5,0),p1=(0,0),p2=(0,n_filas),angle=90,dxfattribs={"layer":"COTAS"}).render()
    doc.saveas(path_out); print(f"  ✅  DXF: {path_out}")

def main():
    print("\n  🌿  HyphaPod — Plano v5 (leyenda lateral)\n")
    for f in [EXCEL_DISEÑO,EXCEL_CATALOGO]:
        if not os.path.exists(f): print(f"  ✗  {f}"); sys.exit(1)
    cat=cargar_catalogo(EXCEL_CATALOGO)
    celdas,n_cols,n_filas=leer_diseño(EXCEL_DISEÑO,cat)
    total=sum(len(v) for v in celdas.values())
    print(f"  Parcela: {n_cols}×{n_filas} m  |  {total} plantas\n")
    generar_pdf(celdas,n_cols,n_filas,cat,os.path.join(_DISEÑOS_DIR,"HyphaPod_Plano_v5.pdf"))
    generar_dxf(celdas,n_cols,n_filas,cat,os.path.join(_DISEÑOS_DIR,"HyphaPod_Plano_v5.dxf"))
    print()

if __name__=="__main__":
    main()
