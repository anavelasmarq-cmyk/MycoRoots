"""
═══════════════════════════════════════════════════════════════════════
  HyphaPod Generator v4 — MycoRoots Network
  Generador de diseño de microbosque urbano (Método Miyawaki)
  Adaptado a clima mediterráneo · Córdoba, España
  TFG Ingeniería Agroalimentaria — Universidad de Córdoba

  CAMBIO PRINCIPAL respecto a v2:
  Cada celda contiene plantas de VARIOS ESTRATOS mezclados, excepto:
    • Anillo 1 (borde exterior) → solo CA (Caméfitas)
    • Celdas con estrato MG     → solo MG, 1 planta (sin mezcla)
  En el resto (anillos 2+, sin MG): 3 plantas/m² de estratos distintos
  elegidos probabilísticamente según el gradiente concéntrico.
═══════════════════════════════════════════════════════════════════════
  Uso: python hyphapod_generator_v4.py
  Requiere: openpyxl, catalogo_especies_v5.xlsx en el mismo directorio
═══════════════════════════════════════════════════════════════════════
"""

import random
import unicodedata
import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────
# CONSTANTES Y CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
CATALOGO_FILE  = os.path.join(SCRIPT_DIR, "catalogo_especies_v5.xlsx")
DISEÑOS_DIR    = os.path.join(SCRIPT_DIR, "..", "diseños")
CATALOGO_SHEET = "CATÁLOGO DE ESPECIES"

PRECIO_MEDIO     = {"ECO": 0.65, "MED": 1.15, "PRE": 2.50}
DENSIDAD_ESTRATO = {"CA": 3, "MI": 2, "ME": 2, "MG": 1}
DENSIDAD_MIXTA   = 3  # plantas/m² en celdas mezcladas (anillos 2+, sin MG)

# Gradiente de estratos por posición normalizada del anillo.
# t=0 → borde exterior  |  t=1 → núcleo central
# Cada fila: (t, peso_CA, peso_MI, peso_ME, peso_MG)
# CA siempre tiene el peso más alto entre los no-MG para favorecer su presencia.
# ME=0 en anillos 1 y 2 (t≤0.25); solo aparece desde anillo 3 hacia el centro.
GRADIENTE_ANILLOS = [
    (0.00, 10,  0,  0,  0),   # Anillo 1: CA solo
    (0.25,  6,  9,  0,  0),   # Anillo 2: MI dominante, CA secundario
    (0.50,  3,  8,  4,  0),   # Anillo 3: MI dominante, ME media-baja
    (0.75,  2,  3, 10,  0),   # Anillo 4+: ME muy dominante
    (1.00,  2,  2,  7, 10),   # Centro: MG dominante (celda propia), ME en MIX
]

COLOR_ESTRATO = {
    "CA": "D6EAD0",
    "MI": "B7D7A8",
    "ME": "7DB77B",
    "MG": "3A6B27",
}
# Un color distinto por combinación de estratos, gradiente claro→oscuro.
COLOR_COMBINACION = {
    frozenset(["CA"]):          "D6EAD0",  # 1 · verde muy claro   — CA solo
    frozenset(["CA", "MI"]):    "BBDCA8",  # 2 · verde claro       — CA + MI
    frozenset(["MI"]):          "9FCD88",  # 3 · verde medio-claro — MI solo
    frozenset(["CA", "ME"]):    "83BE70",  # 4 · verde medio       — CA + ME
    frozenset(["MI", "ME"]):    "68AF58",  # 5 · verde medio-oscuro — MI + ME
    frozenset(["ME"]):          "4E9940",  # 6 · verde oscuro       — ME solo
    frozenset(["MG"]):          "3A6B27",  # 7 · verde muy oscuro   — MG solo
}
HDR_COLOR = "1E3A0E"

# ─────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────

def normalizar(texto):
    """Elimina tildes y convierte a minúsculas para comparación."""
    texto = texto.strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')

def col_letra(n):
    """Convierte número de columna (1-based) a letra Excel: 1→A, 27→AA."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

def input_int(prompt, minv, maxv):
    """Pide un entero al usuario dentro de un rango."""
    while True:
        try:
            val = int(input(prompt))
            if minv <= val <= maxv:
                return val
            print(f"  ⚠  Introduce un número entre {minv} y {maxv}.")
        except ValueError:
            print("  ⚠  Introduce un número entero válido.")

def input_float(prompt, minv=0):
    """Pide un número decimal al usuario."""
    while True:
        try:
            val = float(input(prompt).replace(",", "."))
            if val >= minv:
                return val
            print(f"  ⚠  El valor debe ser mayor que {minv}.")
        except ValueError:
            print("  ⚠  Introduce un número válido (ej: 500 o 500.00).")

def input_yn(prompt):
    """Pregunta sí/no. Devuelve True para sí."""
    while True:
        r = input(prompt + " (s/n): ").strip().lower()
        if r in ("s", "si", "sí", "yes", "y"):
            return True
        if r in ("n", "no"):
            return False
        print("  ⚠  Responde 's' o 'n'.")

def separador():
    print("\n" + "─" * 60)

# ─────────────────────────────────────────────────────────────────────
# CARGA DEL CATÁLOGO
# ─────────────────────────────────────────────────────────────────────

def cargar_catalogo():
    if not os.path.exists(CATALOGO_FILE):
        print(f"\n  ✗ No se encuentra '{CATALOGO_FILE}' en el directorio actual.")
        print("    Asegúrate de que el archivo está junto a este script.")
        sys.exit(1)

    wb = openpyxl.load_workbook(CATALOGO_FILE, data_only=True)
    ws = wb[CATALOGO_SHEET]

    catalogo = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        cod = row[0]
        if not cod or not str(cod).strip():
            continue
        cod = str(cod).strip()
        if cod.startswith("⚙"):
            continue
        try:
            nombre_comun      = str(row[1]).strip()
            nombre_cientifico = str(row[2]).strip()
            posicion          = str(row[20]).strip()
            plantas_m2        = int(row[19])
            cat_precio        = str(row[17]).strip()
            precio_ref        = float(row[18]) if row[18] else PRECIO_MEDIO.get(cat_precio, 0.65)
        except (TypeError, ValueError, IndexError):
            continue

        cod_estrato = str(row[3]).strip() if row[3] else cod[:2]

        catalogo.append({
            "cod":               cod,
            "nombre_comun":      nombre_comun,
            "nombre_norm":       normalizar(nombre_comun),
            "cod_norm":          normalizar(cod),
            "nombre_cientifico": nombre_cientifico,
            "cod_estrato":       cod_estrato,
            "posicion":          posicion,
            "plantas_m2":        plantas_m2,
            "cat_precio":        cat_precio,
            "precio_medio":      precio_ref,
        })

    if not catalogo:
        print("  ✗ El catálogo está vacío o no se pudo leer correctamente.")
        sys.exit(1)

    return catalogo


def buscar_especie(catalogo, texto):
    """Busca una especie por código o nombre común (insensible a tildes/mayúsculas)."""
    t = normalizar(texto)
    for e in catalogo:
        if e["cod_norm"] == t or e["nombre_norm"] == t:
            return e
    return None

# ─────────────────────────────────────────────────────────────────────
# DIÁLOGO CON EL USUARIO
# ─────────────────────────────────────────────────────────────────────

def dialogo(catalogo):
    print("\n")
    print("╔══════════════════════════════════════════════════════════╗")
    print("║    🌿  MycoRoots — Generador de HyphaPod  v3  🌿       ║")
    print("║  Microbosque urbano · Método Miyawaki · Córdoba         ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()
    print("  Bienvenido. Versión 3: celdas con MEZCLA de estratos.")
    print("  Anillo 1 → solo CA  |  Celdas MG → solo MG  |  Resto → mix")
    separador()

    # ── 1. DIMENSIONES ────────────────────────────────────────────────
    print("\n📐  DIMENSIONES DE LA PARCELA")
    print("  La parcela debe ser rectangular, entre 100 y 300 m².")
    print("  Introduce las medidas en metros enteros (ej: 10 x 15).\n")
    while True:
        largo = input_int("  Largo (m) [10–20]: ", 10, 20)
        ancho = input_int("  Ancho (m) [10–20]: ", 10, 20)
        superficie = largo * ancho
        if 100 <= superficie <= 300:
            break
        print(f"  ⚠  La superficie resultante ({superficie} m²) está fuera del rango 100–300 m².")

    n_anillos = (min(ancho, largo) - 1) // 2 + 1
    print(f"\n  ✓ Parcela: {largo} m × {ancho} m = {superficie} m²")
    print(f"  ✓ Anillos concéntricos en este diseño: {n_anillos}")

    # ── 2. PRESUPUESTO ────────────────────────────────────────────────
    separador()
    print("\n💶  PRESUPUESTO")
    print("  Introduce tu presupuesto máximo para la compra de plantas.\n")
    presupuesto = input_float("  Presupuesto máximo (€): ", minv=1)

    # ── 3. TIPO DE DISEÑO ─────────────────────────────────────────────
    separador()
    print("\n🌱  TIPO DE DISEÑO")
    print("  Opción A — Diseño por defecto:")
    print("    Aleatorio con todas las especies del catálogo.")
    print()
    print("  Opción B — Diseño personalizado:")
    print("    Tú eliges estratos, exclusiones y preferencias de diseño.")
    print()
    personalizar = input_yn("  ¿Deseas personalizar tu HyphaPod?")

    estratos_seleccionados = ["CA", "MI", "ME", "MG"]
    especies_excluidas     = []
    fijaciones             = []

    if personalizar:
        # ── 4. ESTRATOS ───────────────────────────────────────────────
        separador()
        print("\n🌳  ESTRATOS A INCLUIR")
        print("  CA → Caméfitas         (<1 m)")
        print("  MI → Microfanerófitas  (1–3 m)")
        print("  ME → Mesofanerófitas   (3–8 m)")
        print("  MG → Megafanerófitas   (>8 m)")
        print()
        print("  Códigos separados por comas, o ENTER para todos.\n")

        while True:
            entrada = input("  Estratos: ").strip()
            if not entrada:
                estratos_seleccionados = ["CA", "MI", "ME", "MG"]
                break
            candidatos = [e.strip().upper() for e in entrada.split(",")]
            invalidos  = [c for c in candidatos if c not in ("CA", "MI", "ME", "MG")]
            if invalidos:
                print(f"  ⚠  Códigos no reconocidos: {invalidos}. Usa CA, MI, ME, MG.")
            else:
                estratos_seleccionados = candidatos
                break

        print(f"\n  ✓ Estratos: {', '.join(estratos_seleccionados)}")

        # ── 5. EXCLUSIONES ────────────────────────────────────────────
        separador()
        print("\n🚫  EXCLUSIONES")
        print("  Código (ej: MG10) o nombre común. Varios separados por comas.")
        print("  ENTER para omitir.\n")

        entrada = input("  Especies a excluir: ").strip()
        if entrada:
            for item in entrada.split(","):
                item = item.strip()
                if not item:
                    continue
                esp = buscar_especie(catalogo, item)
                if esp:
                    especies_excluidas.append(esp["cod"])
                    print(f"  ✓ Excluida: {esp['cod']} — {esp['nombre_comun']}")
                else:
                    print(f"  ⚠  No encontrada: '{item}' — revisa el catálogo.")

        # ── 6. PREFERENCIAS DE DISEÑO ─────────────────────────────────
        separador()
        print("\n📌  PREFERENCIAS DE DISEÑO")
        print(f"  Anillo 1 = borde exterior  |  Anillo {n_anillos} = núcleo central.")
        print("  ENTER para omitir.\n")

        while input_yn("  ¿Añadir una preferencia de diseño?"):
            print()
            entrada_esp = input("  Especie (código o nombre): ").strip()
            esp = buscar_especie(catalogo, entrada_esp)
            if not esp:
                print(f"  ⚠  No encontrada: '{entrada_esp}'. Inténtalo de nuevo.")
                continue

            if esp["cod_estrato"] not in estratos_seleccionados:
                print(f"  ⚠  '{esp['nombre_comun']}' es estrato {esp['cod_estrato']}, "
                      f"no incluido ({', '.join(estratos_seleccionados)}).")
                print(f"     Se añade el estrato {esp['cod_estrato']} al pool activo.")
                estratos_seleccionados.append(esp["cod_estrato"])

            while True:
                zona_op = input(f"  Anillo de preferencia [1–{n_anillos}]: ").strip()
                if zona_op.isdigit() and 1 <= int(zona_op) <= n_anillos:
                    break
                print(f"  ⚠  Introduce un número entre 1 y {n_anillos}.")
            zona = int(zona_op)

            t_natural = {"CA": 0.0, "MI": 0.4, "ME": 0.8, "MG": 1.0}
            anillo_natural = round(t_natural.get(esp["cod_estrato"], 0.5) * (n_anillos - 1)) + 1
            if zona != anillo_natural:
                print(f"  ℹ  Nota ecológica: estrato {esp['cod_estrato']} suele ir en "
                      f"el anillo {anillo_natural}. Se respeta tu elección.")

            print("  ℹ  Orientación de probabilidad (los pesos de posición base van de 0.1 a 10):")
            print("       1–10  → influencia baja, la especie puede aparecer pero no domina")
            print("      25–50  → influencia moderada, compite con el gradiente de posición")
            print("     75–100  → influencia alta, la especie domina claramente en ese anillo")
            prob = input_int("  Probabilidad de aparición en ese anillo [1–100] %: ", 1, 100)

            ya_fijada = any(f["especie"]["cod"] == esp["cod"] and f["zona"] == zona
                            for f in fijaciones)
            if ya_fijada:
                print(f"  ⚠  {esp['cod']} ya tiene preferencia en el anillo {zona}. Se omite el duplicado.")
            else:
                fijaciones.append({"especie": esp, "zona": zona, "prob": prob})
                print(f"  ✓ Preferencia de diseño añadida: {esp['cod']} — {esp['nombre_comun']} → anillo {zona} · {prob}% de aparición")
            print()

    return {
        "largo":       largo,
        "ancho":       ancho,
        "superficie":  superficie,
        "presupuesto": presupuesto,
        "estratos":    estratos_seleccionados,
        "excluidas":   especies_excluidas,
        "preferencias":  fijaciones,
    }

# ─────────────────────────────────────────────────────────────────────
# GRADIENTE PROBABILÍSTICO
# ─────────────────────────────────────────────────────────────────────

def pesos_estrato_anillo(t, estratos_disponibles):
    """
    Interpola los pesos de estrato en GRADIENTE_ANILLOS según t ∈ [0,1].
    Solo devuelve estratos presentes en estratos_disponibles.
    """
    pts = GRADIENTE_ANILLOS
    if t <= pts[0][0]:
        _, ca, mi, me, mg = pts[0]
    elif t >= pts[-1][0]:
        _, ca, mi, me, mg = pts[-1]
    else:
        ca = mi = me = mg = 0.0
        for i in range(len(pts) - 1):
            t0, ca0, mi0, me0, mg0 = pts[i]
            t1, ca1, mi1, me1, mg1 = pts[i + 1]
            if t0 <= t <= t1:
                frac = (t - t0) / (t1 - t0)
                ca = ca0 + frac * (ca1 - ca0)
                mi = mi0 + frac * (mi1 - mi0)
                me = me0 + frac * (me1 - me0)
                mg = mg0 + frac * (mg1 - mg0)
                break

    pesos = {k: v for k, v in {"CA": ca, "MI": mi, "ME": me, "MG": mg}.items()
             if k in estratos_disponibles and v > 0}
    if not pesos:
        pesos = {e: 1.0 for e in estratos_disponibles}
    return pesos


def pesos_posicion_anillo(t):
    """
    Pesos para el campo 'posicion' de especie según t ∈ [0,1].
    t=0 → favorece BORDE  |  t=0.5 → favorece BORDE-INTERIOR  |  t=1 → favorece INTERIOR
    """
    w_borde     = max(0.1, 10.0 * (1.0 - t))
    w_borde_int = max(0.1, 10.0 * (1.0 - abs(2.0 * t - 1.0)))
    w_interior  = max(0.1, 10.0 * t)
    return {"BORDE": w_borde, "BORDE-INTERIOR": w_borde_int, "INTERIOR": w_interior}

# ─────────────────────────────────────────────────────────────────────
# SELECCIÓN DE ESPECIES POR CELDA
# ─────────────────────────────────────────────────────────────────────

def seleccionar_especies_celda(pool_c, t, fijaciones_activas, n_plantas):
    """
    Selecciona n_plantas especies distintas del pool_c (un solo estrato).
    Ponderación por posición preferente según t.
    Las preferencias de diseño activas reciben el peso indicado por el usuario (prob*0.2).
    """
    pesos_pos  = pesos_posicion_anillo(t)
    candidatos = list(pool_c)
    weights    = [max(0.1, pesos_pos.get(esp["posicion"], 1.0)) for esp in candidatos]

    for f in fijaciones_activas:
        esp, peso = f["especie"], f["peso"]
        if esp in candidatos:
            weights[candidatos.index(esp)] = peso
        else:
            candidatos.insert(0, esp)
            weights.insert(0, peso)

    seleccionadas = []
    usados_cod    = set()

    for _ in range(n_plantas):
        disp_c = [c for c in candidatos if c["cod"] not in usados_cod]
        disp_w = [weights[candidatos.index(c)] for c in disp_c]
        if not disp_c:
            disp_c, disp_w = candidatos, weights
        elegida = random.choices(disp_c, weights=disp_w, k=1)[0]
        seleccionadas.append(elegida)
        usados_cod.add(elegida["cod"])

    return seleccionadas


def seleccionar_plantas_mixtas(pool_completo, t, estratos_sin_MG, fij_activas, n_plantas):
    """
    Selecciona n_plantas de estratos MEZCLADOS (sin MG).

    Planta dominante (slot 0): el estrato con MAYOR peso en el gradiente para esta t
    → garantiza el gradiente de altura borde (CA) → centro (ME).
    Plantas acompañantes (slots 1+): pesos uniformes → más aleatorio.

    Reglas adicionales:
      • Máximo 1 planta de estrato CA por celda mixta.
      • Si hay alguna planta de estrato ME, el máximo de plantas en la celda es 2.
    """
    pesos_est = pesos_estrato_anillo(t, estratos_sin_MG)
    pesos_pos = pesos_posicion_anillo(t)

    plantas         = []
    usados          = set()
    n_CA_en_celda   = 0
    hay_ME_en_celda = False

    def _filtrar(pesos_base):
        """Aplica restricciones CA-máx-1 y ME-cap-2 sobre pesos_base."""
        f = {
            est: w for est, w in pesos_base.items()
            if not (est == "CA" and n_CA_en_celda >= 1)
            and not (est == "ME" and len(plantas) >= 2)
        }
        if not f:
            f = {est: w for est, w in pesos_base.items()
                 if not (est == "ME" and len(plantas) >= 2)}
        return f or dict(pesos_base)

    for slot in range(n_plantas):
        if hay_ME_en_celda and len(plantas) >= 2:
            break

        pesos_actuales = _filtrar(pesos_est)
        if not pesos_actuales:
            break

        ests = list(pesos_actuales.keys())
        ws_e = list(pesos_actuales.values())

        if slot == 0:
            # Planta dominante: estrato con mayor peso → gradiente garantizado
            estrato = ests[ws_e.index(max(ws_e))]
        else:
            # Plantas acompañantes: pesos uniformes → aleatorio entre disponibles
            pesos_unif = _filtrar({est: 1.0 for est in pesos_est})
            estrato = random.choices(list(pesos_unif.keys()),
                                     weights=list(pesos_unif.values()), k=1)[0]

        pool_e = [e for e in pool_completo
                  if e["cod_estrato"] == estrato and e["cod"] not in usados]
        if not pool_e:
            pool_e = [e for e in pool_completo if e["cod_estrato"] == estrato]
        if not pool_e:
            continue

        pesos_w = [max(0.1, pesos_pos.get(e["posicion"], 1.0)) for e in pool_e]

        fij_est = [f for f in fij_activas if f["especie"]["cod_estrato"] == estrato]
        for f in fij_est:
            esp, peso = f["especie"], f["peso"]
            if esp in pool_e:
                pesos_w[pool_e.index(esp)] = peso
            else:
                pool_e.insert(0, esp)
                pesos_w.insert(0, peso)

        esp = random.choices(pool_e, weights=pesos_w, k=1)[0]
        plantas.append(esp)
        usados.add(esp["cod"])

        if estrato == "CA":
            n_CA_en_celda += 1
        if estrato == "ME":
            hay_ME_en_celda = True

    return plantas

# ─────────────────────────────────────────────────────────────────────
# MOTOR DE DISEÑO
# ─────────────────────────────────────────────────────────────────────

def generar_grid(params, catalogo):
    """
    Genera la cuadrícula del HyphaPod v3 con celdas de estratos mezclados.

    Reglas por celda:
      • Anillo 1 (dist=0) → solo CA, 3 plantas/m²
      • Sorteado MG       → solo MG, 1 planta/m² (sin mezcla, celda exclusiva)
      • Resto (MIX)       → mezcla de estratos sin MG (CA+MI, CA+ME, MI+ME…):
          - CA tiene la mayor probabilidad entre los no-MG
          - Máx 1 CA por celda mixta
          - Si la celda incluye ME → máx 2 plantas totales; si no → 3
    """
    n_filas  = params["ancho"]
    n_cols   = params["largo"]
    max_dist = (min(n_filas, n_cols) - 1) // 2

    pool_completo = [
        e for e in catalogo
        if e["cod_estrato"] in params["estratos"]
        and e["cod"] not in params["excluidas"]
    ]
    if not pool_completo:
        return None, "No hay especies disponibles con los filtros seleccionados."

    estratos_disponibles = {e["cod_estrato"] for e in pool_completo}
    estratos_sin_MG      = estratos_disponibles - {"MG"}

    grid = []
    for f in range(n_filas):
        fila = []
        for c in range(n_cols):
            dist       = min(f, c, n_filas - 1 - f, n_cols - 1 - c)
            t          = dist / max(max_dist, 1)  # 0.0 = borde exterior, 1.0 = núcleo
            anillo_num = dist + 1

            # Preferencias de diseño activas en este anillo (con su peso = prob*0.2)
            fij_anillo = [{"especie": f_["especie"], "peso": f_["prob"] * 0.2}
                          for f_ in params["preferencias"] if f_["zona"] == anillo_num]

            # ── REGLA 1: Anillo 1 → solo CA (+ preferencias de diseño de cualquier estrato)
            if dist == 0:
                pool_CA = [e for e in pool_completo if e["cod_estrato"] == "CA"]
                if not pool_CA:
                    pool_CA = pool_completo
                # Incluir todas las preferencias de diseño del anillo 1
                fij_CA = fij_anillo
                for pref in fij_CA:
                    if pref["especie"] not in pool_CA:
                        pool_CA = list(pool_CA) + [pref["especie"]]
                n_pl    = DENSIDAD_ESTRATO["CA"]
                especies = seleccionar_especies_celda(pool_CA, t, fij_CA, n_pl)
                fila.append({
                    "especies":  especies,
                    "anillo":    dist,
                    "t":         t,
                    "n_plantas": len(especies),
                    "tipo":      "CA",
                })
                continue

            # ── REGLA 2: Sorteo de tipo de celda ─────────────────────
            pesos_est    = pesos_estrato_anillo(t, estratos_disponibles)
            est_sorteado = random.choices(
                list(pesos_est.keys()), weights=list(pesos_est.values()), k=1
            )[0]

            if est_sorteado == "MG":
                # MG solo, 1 planta
                pool_MG  = [e for e in pool_completo if e["cod_estrato"] == "MG"]
                if not pool_MG:
                    pool_MG = pool_completo
                fij_MG   = [f for f in fij_anillo if f["especie"]["cod_estrato"] == "MG"]
                especies = seleccionar_especies_celda(pool_MG, t, fij_MG, 1)
                fila.append({
                    "especies":  especies,
                    "anillo":    dist,
                    "t":         t,
                    "n_plantas": len(especies),
                    "tipo":      "MG",
                })
            else:
                # ── REGLA 3: Mezcla de estratos (sin MG) ─────────────
                est_disp_mix = estratos_sin_MG if estratos_sin_MG else estratos_disponibles

                # ME no puede aparecer en anillos 1 y 2 salvo preferencia de diseño explícita
                if anillo_num <= 2:
                    fij_ME_aqui = [f for f in fij_anillo if f["especie"]["cod_estrato"] == "ME"]
                    if not fij_ME_aqui:
                        est_disp_mix = est_disp_mix - {"ME"}
                        if not est_disp_mix:
                            est_disp_mix = estratos_sin_MG

                especies = seleccionar_plantas_mixtas(
                    pool_completo, t, est_disp_mix, fij_anillo, DENSIDAD_MIXTA
                )
                fila.append({
                    "especies":  especies,
                    "anillo":    dist,
                    "t":         t,
                    "n_plantas": len(especies),
                    "tipo":      "MIX",
                })

        grid.append(fila)

    return grid, None

# ─────────────────────────────────────────────────────────────────────
# CÁLCULO DE COSTE
# ─────────────────────────────────────────────────────────────────────

def calcular_coste(grid):
    """Devuelve coste total y desglose por especie."""
    desglose      = {}
    total_plantas = 0
    total_coste   = 0.0

    for fila in grid:
        for celda in fila:
            for esp in celda["especies"]:
                cod = esp["cod"]
                if cod not in desglose:
                    desglose[cod] = {
                        "nombre":    esp["nombre_comun"],
                        "n_plantas": 0,
                        "precio_ud": esp["precio_medio"],
                        "cat":       esp["cat_precio"],
                        "subtotal":  0.0,
                    }
                desglose[cod]["n_plantas"] += 1
                desglose[cod]["subtotal"]  += esp["precio_medio"]
                total_plantas += 1
                total_coste   += esp["precio_medio"]

    return total_coste, total_plantas, desglose

# ─────────────────────────────────────────────────────────────────────
# GESTIÓN DE PRESUPUESTO
# ─────────────────────────────────────────────────────────────────────

def gestionar_presupuesto(coste, presupuesto, params, catalogo):
    """Si el coste supera el presupuesto, ofrece tres opciones."""
    if coste <= presupuesto:
        return params, False

    separador()
    sobrecoste = coste - presupuesto
    print(f"\n  ⚠  El diseño cuesta {coste:.2f} € y supera el presupuesto")
    print(f"     de {presupuesto:.2f} € en {sobrecoste:.2f} €.\n")
    print("  A — Reducir la superficie (nuevas dimensiones)")
    print("  B — Excluir las especies más caras (PRE y/o MED)")
    print("  C — Conservar el diseño aunque supere el presupuesto")
    print()

    while True:
        op = input("  Elige opción [A/B/C]: ").strip().upper()
        if op == "A":
            print("\n  Introduce las nuevas dimensiones:")
            while True:
                largo = input_int("  Largo (m) [10–20]: ", 10, 20)
                ancho = input_int("  Ancho (m) [10–20]: ", 10, 20)
                sup   = largo * ancho
                if 100 <= sup <= 300:
                    break
                print(f"  ⚠  Superficie {sup} m² fuera del rango 100–300 m².")
            params = dict(params)
            params["largo"]      = largo
            params["ancho"]      = ancho
            params["superficie"] = sup
            return params, True

        elif op == "B":
            print("\n  ¿Excluir categoría PRE, MED o ambas?")
            print("  1 — Solo PRE")
            print("  2 — PRE y MED (solo quedan ECO)")
            while True:
                sub = input("  Elige [1/2]: ").strip()
                if sub in ("1", "2"):
                    break
                print("  ⚠  Introduce 1 o 2.")
            cats_excluir     = ["PRE"] if sub == "1" else ["PRE", "MED"]
            nuevas_excluidas = list(params["excluidas"])
            for esp in catalogo:
                if esp["cat_precio"] in cats_excluir and esp["cod"] not in nuevas_excluidas:
                    nuevas_excluidas.append(esp["cod"])
            params = dict(params)
            params["excluidas"] = nuevas_excluidas
            print(f"\n  ✓ Excluidas las especies de categoría {cats_excluir}.")
            return params, True

        elif op == "C":
            print(f"\n  ✓ Se exportará el diseño con coste {coste:.2f} €.")
            return params, False

        else:
            print("  ⚠  Introduce A, B o C.")

# ─────────────────────────────────────────────────────────────────────
# EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────────────────────────────

def _color_celda(celda):
    """Devuelve (bg, fg) para una celda según la combinación exacta de estratos presentes."""
    estratos = frozenset(esp["cod_estrato"] for esp in celda["especies"])
    bg = COLOR_COMBINACION.get(estratos, "D9D9D9")   # gris como fallback visible
    fg = "FFFFFF" if estratos == frozenset(["MG"]) else "111111"
    return bg, fg


def exportar_excel(grid, params, coste, total_plantas, desglose, nombre_archivo, n_anillos):
    """Genera el Excel con la cuadrícula del HyphaPod y el resumen económico."""
    wb   = Workbook()
    thin = Side(style="thin", color="AAAAAA")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sty(c, bg=None, fg="111111", bold=False, size=9, ah="center", wrap=True):
        c.font = Font(bold=bold, color=fg, name="Arial", size=size)
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=ah, vertical="center", wrap_text=wrap)
        c.border = brd

    n_filas  = params["ancho"]
    n_cols   = params["largo"]
    last_col = col_letra(n_cols + 1)

    # ── HOJA 1: CUADRÍCULA ────────────────────────────────────────────
    ws = wb.active
    ws.title = "CUADRÍCULA HyphaPod"
    ws.freeze_panes = "B2"

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = (f"HyphaPod v3 — MycoRoots  |  {params['largo']} m × {params['ancho']} m "
               f"= {params['superficie']} m²  |  {n_anillos} anillos  |  "
               f"{total_plantas} plantas  |  Coste estimado: {coste:.2f} €")
    sty(c, bg=HDR_COLOR, fg="FFFFFF", bold=True, size=11)
    ws.row_dimensions[1].height = 22

    ws.column_dimensions["A"].width = 5
    for c_idx in range(n_cols):
        col_str = col_letra(c_idx + 2)
        ws.column_dimensions[col_str].width = 22
        cell = ws.cell(row=2, column=c_idx + 2)
        cell.value = col_letra(c_idx + 1)
        sty(cell, bg="D9D9D9", bold=True, size=9)
    ws.row_dimensions[2].height = 16

    for f_idx, fila in enumerate(grid):
        row_excel = f_idx + 3
        ws.row_dimensions[row_excel].height = 52

        cell_num = ws.cell(row=row_excel, column=1)
        cell_num.value = f_idx + 1
        sty(cell_num, bg="D9D9D9", bold=True, size=9)

        for c_idx, celda in enumerate(fila):
            col_excel = c_idx + 2
            especies  = celda["especies"]
            anillo    = celda["anillo"]
            n         = celda["n_plantas"]

            bg_celda, fg_celda = _color_celda(celda)

            lineas    = [f"{esp['cod']} {esp['nombre_comun']}" for esp in especies]
            contenido = "\n".join(lineas) + f"\nA{anillo + 1} · {n} pl./m²"

            cell = ws.cell(row=row_excel, column=col_excel)
            cell.value = contenido
            sty(cell, bg=bg_celda, fg=fg_celda, size=7, wrap=True)

    # Leyenda
    leyenda_row = n_filas + 4
    ws.merge_cells(f"A{leyenda_row}:{last_col}{leyenda_row}")
    lc = ws[f"A{leyenda_row}"]
    lc.value = (f"LEYENDA DE CELDAS  "
                f"(diseño con {n_anillos} anillos · v3 mezcla de estratos · CA/MI/ME/MG)")
    sty(lc, bg=HDR_COLOR, fg="FFFFFF", bold=True, ah="left")

    grad_items = [
        ("CA  solo",
         "Solo Caméfitas (<1 m) · anillo 1 (borde) · 3 pl./m²",
         COLOR_COMBINACION[frozenset(["CA"])], "111111"),
        ("CA + MI",
         "Mezcla Caméfitas + Microfanerófitas · anillos 2-3 · máx 3 pl./m² · 1 CA máx",
         COLOR_COMBINACION[frozenset(["CA","MI"])], "111111"),
        ("MI  solo",
         "Solo Microfanerófitas (1–3 m) · anillos 2-3",
         COLOR_COMBINACION[frozenset(["MI"])], "111111"),
        ("CA + ME",
         "Mezcla Caméfitas + Mesofanerófitas · anillos 3+ · máx 2 pl./m² · 1 CA máx",
         COLOR_COMBINACION[frozenset(["CA","ME"])], "111111"),
        ("MI + ME",
         "Mezcla Microfanerófitas + Mesofanerófitas · anillos 3+ · máx 2 pl./m²",
         COLOR_COMBINACION[frozenset(["MI","ME"])], "111111"),
        ("ME  solo",
         "Solo Mesofanerófitas (3–8 m) · anillos 3+ · máx 2 pl./m²",
         COLOR_COMBINACION[frozenset(["ME"])], "111111"),
        ("MG  Núcleo",
         "Solo Megafanerófitas (>8 m) · celda exclusiva · 1 pl./m²",
         COLOR_COMBINACION[frozenset(["MG"])], "FFFFFF"),
    ]
    for i, (nombre, desc, color, fg_l) in enumerate(grad_items):
        r  = leyenda_row + 1 + i
        c1 = ws.cell(row=r, column=1)
        c1.value = nombre
        sty(c1, bg=color, fg=fg_l, bold=True, size=8)
        ws.merge_cells(f"B{r}:D{r}")
        c2 = ws.cell(row=r, column=2)
        c2.value = desc
        sty(c2, bg=color, fg=fg_l, size=8, ah="left")

    # ── HOJA 2: RESUMEN ECONÓMICO ─────────────────────────────────────
    ws2 = wb.create_sheet("RESUMEN ECONÓMICO")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 14

    ws2.merge_cells("A1:E1")
    c = ws2["A1"]
    c.value = "RESUMEN ECONÓMICO — HyphaPod v3 · MycoRoots"
    sty(c, bg=HDR_COLOR, fg="FFFFFF", bold=True, size=12)
    ws2.row_dimensions[1].height = 22

    hdrs = ["CÓD.", "ESPECIE", "N.º PLANTAS", "PRECIO/UD (€)", "SUBTOTAL (€)"]
    for j, h in enumerate(hdrs):
        cell = ws2.cell(row=2, column=j + 1)
        cell.value = h
        sty(cell, bg="4E8E3A", fg="FFFFFF", bold=True, size=9)
    ws2.row_dimensions[2].height = 18

    for i, (cod, d) in enumerate(sorted(desglose.items())):
        r  = i + 3
        bg = COLOR_ESTRATO.get("MG" if cod.startswith("MG") else cod[:2], "FFFFFF")
        fg = "FFFFFF" if cod.startswith("MG") else "111111"
        vals = [cod, d["nombre"], d["n_plantas"],
                f"{d['precio_ud']:.2f}", f"{d['subtotal']:.2f}"]
        for j, v in enumerate(vals):
            cell = ws2.cell(row=r, column=j + 1)
            cell.value = v
            sty(cell, bg=bg, fg=fg, size=9, ah="center" if j != 1 else "left")
        ws2.row_dimensions[r].height = 16

    total_row = len(desglose) + 3
    ws2.merge_cells(f"A{total_row}:D{total_row}")
    ct = ws2[f"A{total_row}"]
    ct.value = f"TOTAL — {total_plantas} plantas"
    sty(ct, bg=HDR_COLOR, fg="FFFFFF", bold=True, size=10, ah="right")
    cv = ws2[f"E{total_row}"]
    cv.value = f"{coste:.2f} €"
    sty(cv, bg=HDR_COLOR, fg="FFFFFF", bold=True, size=10)
    ws2.row_dimensions[total_row].height = 20

    nota_row = total_row + 2
    ws2.merge_cells(f"A{nota_row}:E{nota_row}")
    cn = ws2[f"A{nota_row}"]
    cn.value = ("⚠  Precios orientativos pendientes de verificación. "
                "Fuentes de referencia consultadas: Viveros Santa Bárbara, Viveros Sierra Norte, "
                "Control Bio, Planflor, Plantas Xerófitas, The Original Garden, Plantamus, "
                "Nina Seeds, Oreshka-Seeds. Verificar antes de presupuestar.")
    cn.font = Font(italic=True, color="666666", name="Arial", size=8)
    cn.fill = PatternFill("solid", start_color="FFF2CC")
    cn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cn.border = brd
    ws2.row_dimensions[nota_row].height = 32

    wb.save(nombre_archivo)

# ─────────────────────────────────────────────────────────────────────
# PROGRAMA PRINCIPAL
# ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(DISEÑOS_DIR, exist_ok=True)
    catalogo = cargar_catalogo()
    params   = dialogo(catalogo)

    max_intentos = 5
    for intento in range(max_intentos):
        separador()
        suffix = f" (intento {intento + 1})" if intento > 0 else ""
        print(f"\n  🔄  Generando diseño{suffix}...")

        grid, error = generar_grid(params, catalogo)

        if error:
            print(f"\n  ⚠  {error}")
            print("     Generando con todos los estratos como fallback...")
            params_fb = dict(params)
            params_fb["estratos"] = ["CA", "MI", "ME", "MG"]
            grid, error2 = generar_grid(params_fb, catalogo)
            if error2:
                print("     Eliminando también exclusiones de especie como último recurso...")
                params_fb["excluidas"] = []
                grid, error2 = generar_grid(params_fb, catalogo)
            if error2:
                print(f"  ✗  Error crítico: {error2}")
                sys.exit(1)

        coste, total_plantas, desglose = calcular_coste(grid)
        params, regenerar = gestionar_presupuesto(coste, params["presupuesto"], params, catalogo)
        if not regenerar:
            break
    else:
        print("\n  ⚠  No se pudo ajustar al presupuesto tras varios intentos.")
        print("     Se exporta el último diseño con indicación del sobrecoste.")

    n_anillos = (min(params["ancho"], params["largo"]) - 1) // 2 + 1

    separador()
    print(f"\n  ✅  DISEÑO GENERADO")
    print(f"      Superficie:      {params['largo']} m × {params['ancho']} m = {params['superficie']} m²")
    print(f"      Anillos:         {n_anillos}")
    print(f"      Total plantas:   {total_plantas}")
    print(f"      Coste estimado:  {coste:.2f} €")
    print(f"      Presupuesto máx: {params['presupuesto']:.2f} €")
    if coste > params["presupuesto"]:
        print(f"      ⚠  Sobrecoste:   {coste - params['presupuesto']:.2f} €")
    print()

    nombre_defecto = f"HyphaPod_v3_{params['largo']}x{params['ancho']}.xlsx"
    print(f"\n  El nombre por defecto del archivo es: {nombre_defecto}")
    if input_yn("  ¿Deseas personalizar el nombre del archivo?"):
        while True:
            nombre_custom = input("  Nombre del archivo (sin extensión): ").strip()
            if nombre_custom:
                nombre_archivo = os.path.join(DISEÑOS_DIR, nombre_custom + ".xlsx")
                break
            print("  ⚠  El nombre no puede estar vacío.")
    else:
        nombre_archivo = os.path.join(DISEÑOS_DIR, nombre_defecto)
    while True:
        try:
            exportar_excel(grid, params, coste, total_plantas, desglose, nombre_archivo, n_anillos)
            break
        except PermissionError:
            print(f"\n  ⚠  No se puede guardar '{nombre_archivo}'.")
            print("     El archivo está abierto en Excel. Ciérralo y pulsa ENTER para reintentar,")
            print("     o escribe un nombre nuevo (sin extensión) para guardar en otro archivo.")
            nuevo = input("  Nuevo nombre o ENTER para reintentar: ").strip()
            if nuevo:
                nombre_archivo = nuevo + ".xlsx"

    separador()
    print(f"\n  📄  Excel exportado: {nombre_archivo}")
    print("      Ábrelo para ver la cuadrícula y el resumen económico.")

    # ── VISUALIZACIONES OPCIONALES ────────────────────────────────────
    separador()
    print("\n  🗺️   VISUALIZACIONES DEL DISEÑO")
    print("      (los scripts deben estar en scripts/ dentro de MycoRoots/)")
    print()

    if input_yn("  ¿Generar plano 2D (PDF + DXF) del diseño?"):
        import shutil, subprocess

        def _instalar_si_falta(paquete):
            try:
                __import__(paquete)
            except ImportError:
                print(f"  📦  Instalando {paquete} en el Python actual...")
                r = subprocess.run([sys.executable, "-m", "pip", "install", paquete])
                if r.returncode != 0:
                    print(f"  ⚠  No se pudo instalar {paquete} automáticamente.")
                    print(f"     Instálalo manualmente desde la consola de Spyder:")
                    print(f"       import subprocess, sys")
                    print(f"       subprocess.run([sys.executable, '-m', 'pip', 'install', '{paquete}'])")

        for pkg in ("ezdxf", "reportlab"):
            _instalar_si_falta(pkg)

        nombre_diseño = os.path.join(DISEÑOS_DIR, "HyphaPod_diseño.xlsx")
        try:
            shutil.copy2(nombre_archivo, nombre_diseño)
            print(f"\n  ✓ Copia guardada como: {nombre_diseño}")
        except Exception as e:
            print(f"\n  ⚠  No se pudo copiar el archivo: {e}")

        _plano_script = os.path.join(SCRIPT_DIR, "hyphapod_plano_v5.py")
        if os.path.exists(_plano_script):
            print("\n  🔄  Generando plano 2D...")
            env_utf8 = {**os.environ, "PYTHONIOENCODING": "utf-8"}
            result = subprocess.run([sys.executable, _plano_script], capture_output=True, text=True, encoding="utf-8", env=env_utf8)
            print(result.stdout.strip())
            if result.returncode != 0:
                print(f"  ⚠  {result.stderr.strip()}")
        else:
            print("  ⚠  No se encontró hyphapod_plano_v5.py — omitido.")

    if input_yn("\n  ¿Generar modelo 3D volumétrico (.glb)?"):
        import shutil, subprocess

        nombre_diseño = os.path.join(DISEÑOS_DIR, "HyphaPod_diseño.xlsx")
        if not os.path.exists(nombre_diseño):
            try:
                shutil.copy2(nombre_archivo, nombre_diseño)
            except Exception as e:
                print(f"  ⚠  No se pudo copiar el archivo: {e}")

        _glb_script = os.path.join(SCRIPT_DIR, "hyphapod_glb.py")
        if os.path.exists(_glb_script):
            print("\n  🔄  Generando modelo 3D...")
            env_utf8 = {**os.environ, "PYTHONIOENCODING": "utf-8"}
            result = subprocess.run([sys.executable, _glb_script], capture_output=True, text=True, encoding="utf-8", env=env_utf8)
            print(result.stdout.strip())
            if result.returncode != 0:
                print(f"  ⚠  {result.stderr.strip()}")
        else:
            print("  ⚠  No se encontró hyphapod_glb.py — omitido.")

    separador()
    print("\n  Gracias por usar MycoRoots 🌿\n")


if __name__ == "__main__":
    main()
