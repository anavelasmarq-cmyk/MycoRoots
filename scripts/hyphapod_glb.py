"""
HyphaPod 3D — exportador GLB (binario, un solo archivo, colores incluidos)
Lee PRUEBIÑA.xlsx + catalogo_especies_v5.xlsx y genera HyphaPod_3D.glb
"""

import math, random, re, os, sys, struct, json
import openpyxl

_SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
_DISEÑOS_DIR   = os.path.join(_SCRIPT_DIR, "..", "diseños")
EXCEL_DISEÑO   = os.path.join(_DISEÑOS_DIR, "HyphaPod_diseño.xlsx")
EXCEL_CATALOGO = os.path.join(_SCRIPT_DIR, "catalogo_especies_v5.xlsx")
HOJA_DISEÑO    = "CUADRÍCULA HyphaPod"
HOJA_CATALOGO  = "CATÁLOGO DE ESPECIES"

# Colores RGBA por estrato (0..1)
COLOR_COPA = {
    "CA": [0.65, 0.88, 0.55, 1.0],
    "MI": [0.28, 0.68, 0.28, 1.0],
    "ME": [0.10, 0.48, 0.18, 1.0],
    "MG": [0.04, 0.26, 0.08, 1.0],
}
COLOR_TRONCO_BASE = [0.38, 0.24, 0.12, 1.0]
COLOR_SUELO       = [0.62, 0.52, 0.36, 1.0]

COPA_RATIO = {"CA": 0.55, "MI": 0.38, "ME": 0.30, "MG": 0.25}
SEG = 7

# ── Catálogo ──────────────────────────────────────────────────────────

def cargar_catalogo(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[HOJA_CATALOGO]
    cat = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        cod = row[0]
        if not cod or str(cod).startswith("⚙"): continue
        cod = str(cod).strip()
        estrato = str(row[3]).strip() if row[3] else cod[:2]
        try:
            alt_min = float(str(row[5]).replace(",","."))
            alt_max = float(str(row[6]).replace(",","."))
        except (ValueError, TypeError):
            alt_min,alt_max = {"CA":(0.3,1.0),"MI":(1.0,3.0),"ME":(3.0,8.0),"MG":(8.0,20.0)}.get(estrato,(1,5))
        cat[cod] = {"estrato":estrato,"alt_min":alt_min,"alt_max":alt_max}
    return cat

# ── Diseño ────────────────────────────────────────────────────────────

def extraer_codigos(texto):
    return re.findall(r'\b(CA\d{2}|MI\d{2}|ME\d{2}|MG\d{2})\b', str(texto)) if texto else []

def leer_diseño(path, catalogo):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[HOJA_DISEÑO]
    datos=[]; n_filas=0; n_cols=0
    for ri, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None: break
        try: int(str(row[0]).strip())
        except: break
        n_filas += 1
        for ci, celda in enumerate(row[1:]):
            if celda is None: continue
            cods = [c for c in extraer_codigos(celda) if c in catalogo]
            if cods:
                datos.append((ci, ri, cods))
                n_cols = max(n_cols, ci+1)
    return datos, n_cols, n_filas

# ── Geometría ─────────────────────────────────────────────────────────

def cilindro_verts(cx, cy, cz, r, h, seg=SEG):
    verts=[]; norms=[]; tris=[]
    base=[]; top=[]
    for i in range(seg):
        a = 2*math.pi*i/seg
        x = cx+r*math.cos(a); z = cz+r*math.sin(a)
        base.append(len(verts)); verts.append([x,cy,z])
        norms.append([math.cos(a),0,math.sin(a)])
        top .append(len(verts)); verts.append([x,cy+h,z])
        norms.append([math.cos(a),0,math.sin(a)])
    for i in range(seg):
        j=(i+1)%seg
        b0,t0=base[i],top[i]; b1,t1=base[j],top[j]
        tris+=[[b0,b1,t1],[b0,t1,t0]]
    # tapa top
    tc=len(verts); verts.append([cx,cy+h,cz]); norms.append([0,1,0])
    for i in range(seg):
        tris.append([tc,top[(i+1)%seg],top[i]])
    return verts,norms,tris

def esfera_verts(cx, cy, cz, r, seg=SEG):
    verts=[]; norms=[]; tris=[]
    rings=max(seg//2,3)
    idx=[]
    for ri in range(rings+1):
        lat=math.pi*ri/rings-math.pi/2
        y=cy+r*math.sin(lat); row=[]
        for si in range(seg):
            lon=2*math.pi*si/seg
            nx=math.cos(lat)*math.cos(lon)
            ny=math.sin(lat)
            nz=math.cos(lat)*math.sin(lon)
            row.append(len(verts))
            verts.append([cx+r*nx,y,cz+r*nz])
            norms.append([nx,ny,nz])
        idx.append(row)
    for ri in range(rings):
        for si in range(seg):
            sn=(si+1)%seg
            v0,v1,v2,v3=idx[ri][si],idx[ri][sn],idx[ri+1][sn],idx[ri+1][si]
            tris+=[[v0,v1,v2],[v0,v2,v3]]
    return verts,norms,tris

def suelo_verts(lx, lz):
    verts=[[0,0,0],[lx,0,0],[lx,0,lz],[0,0,lz]]
    norms=[[0,1,0]]*4
    tris=[[0,1,2],[0,2,3]]
    return verts,norms,tris

# ── GLB builder ───────────────────────────────────────────────────────

def pack_f32(lst):
    return struct.pack(f"{len(lst)}f", *lst)

def pack_u32(lst):
    return struct.pack(f"{len(lst)}I", *lst)

def align4(data):
    r=len(data)%4
    return data+(b'\x00'*(4-r) if r else b'')

def build_glb(meshes):
    """
    meshes: list of {"verts":[[x,y,z]...],"norms":[[nx,ny,nz]...],"tris":[[a,b,c]...],"color":[r,g,b,a]}
    """
    buffers=[]; buffer_views=[]; accessors=[]; m_nodes=[]; materials=[]
    bin_data=b""

    for mi, mesh in enumerate(meshes):
        verts=mesh["verts"]; norms=mesh["norms"]; tris=mesh["tris"]; color=mesh["color"]
        if not verts or not tris: continue

        # Material
        mat_idx=len(materials)
        materials.append({"pbrMetallicRoughness":{
            "baseColorFactor":color,
            "metallicFactor":0.0,"roughnessFactor":0.85},"doubleSided":True})

        # Posiciones
        flat_v=[c for v in verts for c in v]
        min_v=[min(verts,key=lambda x:x[i])[i] for i in range(3)]
        max_v=[max(verts,key=lambda x:x[i])[i] for i in range(3)]
        vdata=pack_f32(flat_v)
        bv_v={"buffer":0,"byteOffset":len(bin_data),"byteLength":len(vdata)}
        buffer_views.append(bv_v); bin_data+=align4(vdata)
        acc_v={"bufferView":len(buffer_views)-1,"componentType":5126,"count":len(verts),
               "type":"VEC3","min":min_v,"max":max_v}
        accessors.append(acc_v); va=len(accessors)-1

        # Normales
        flat_n=[c for n in norms for c in n]
        ndata=pack_f32(flat_n)
        bv_n={"buffer":0,"byteOffset":len(bin_data),"byteLength":len(ndata)}
        buffer_views.append(bv_n); bin_data+=align4(ndata)
        acc_n={"bufferView":len(buffer_views)-1,"componentType":5126,"count":len(norms),"type":"VEC3"}
        accessors.append(acc_n); na=len(accessors)-1

        # Índices
        flat_i=[c for t in tris for c in t]
        idata=pack_u32(flat_i)
        bv_i={"buffer":0,"byteOffset":len(bin_data),"byteLength":len(idata),"target":34963}
        buffer_views.append(bv_i); bin_data+=align4(idata)
        acc_i={"bufferView":len(buffer_views)-1,"componentType":5125,"count":len(flat_i),"type":"SCALAR"}
        accessors.append(acc_i); ia=len(accessors)-1

        mesh_idx=len(m_nodes)
        m_nodes.append({
            "primitives":[{"attributes":{"POSITION":va,"NORMAL":na},"indices":ia,"material":mat_idx}]
        })

    gltf={
        "asset":{"version":"2.0","generator":"HyphaPod-3D"},
        "scene":0,
        "scenes":[{"nodes":list(range(len(m_nodes)))}],
        "nodes":[{"mesh":i} for i in range(len(m_nodes))],
        "meshes":m_nodes,
        "materials":materials,
        "accessors":accessors,
        "bufferViews":buffer_views,
        "buffers":[{"byteLength":len(bin_data)}]
    }

    json_bytes=json.dumps(gltf,separators=(',',':')).encode()
    json_bytes+=b' '*(((len(json_bytes)+3)//4)*4-len(json_bytes))

    bin_bytes=bin_data
    bin_bytes+=b'\x00'*(((len(bin_bytes)+3)//4)*4-len(bin_bytes))

    total=12+8+len(json_bytes)+8+len(bin_bytes)
    glb=struct.pack('<III',0x46546C67,2,total)
    glb+=struct.pack('<II',len(json_bytes),0x4E4F534A)+json_bytes
    glb+=struct.pack('<II',len(bin_bytes), 0x004E4942)+bin_bytes
    return glb

# ── API para uso como librería ─────────────────────────────────────────

def generar_glb_bytes(celdas, n_cols, n_filas, catalogo):
    """
    celdas: dict {(col, fila): [cod1, cod2, ...]}
    catalogo: dict cod → {estrato, alt_min, alt_max}  (salida de cargar_catalogo)
    Devuelve bytes del archivo GLB.
    """
    import random as _rnd
    rng = _rnd.Random(42)
    meshes = []
    sv, sn, st2 = suelo_verts(n_cols, n_filas)
    meshes.append({"verts": sv, "norms": sn, "tris": st2, "color": COLOR_SUELO})
    for (col, fila), cods in celdas.items():
        n_pl = len(cods)
        for i, cod in enumerate(cods):
            if cod not in catalogo:
                continue
            esp = catalogo[cod]
            est = esp["estrato"]
            alt = rng.uniform(esp["alt_min"], esp["alt_max"])
            r_copa = alt * COPA_RATIO.get(est, 0.30) * rng.uniform(0.88, 1.12)
            h_tr = alt * rng.uniform(0.35, 0.42)
            r_tr = max(0.02, r_copa * 0.09)
            if n_pl == 1:
                ox, oz = rng.uniform(-.22, .22), rng.uniform(-.22, .22)
            elif n_pl == 2:
                a = math.pi * i + rng.uniform(-.2, .2)
                ox, oz = math.cos(a) * .26, math.sin(a) * .26
            else:
                a = 2 * math.pi * i / 3 + rng.uniform(-.15, .15)
                ox, oz = math.cos(a) * .26, math.sin(a) * .26
            px, pz = col + 0.5 + ox, fila + 0.5 + oz
            tv, tn, tt = cilindro_verts(px, 0, pz, r_tr, h_tr)
            meshes.append({"verts": tv, "norms": tn, "tris": tt, "color": COLOR_TRONCO_BASE})
            ev, en, et = esfera_verts(px, h_tr + r_copa * 0.55, pz, r_copa)
            meshes.append({"verts": ev, "norms": en, "tris": et, "color": COLOR_COPA[est]})
    return build_glb(meshes)

# ── Main ──────────────────────────────────────────────────────────────

def main():
    print("\n  🌿  HyphaPod 3D GLB — MycoRoots\n")
    for f in [EXCEL_DISEÑO, EXCEL_CATALOGO]:
        if not os.path.exists(f):
            print(f"  ✗  No encontrado: {f}"); sys.exit(1)

    catalogo = cargar_catalogo(EXCEL_CATALOGO)
    datos, n_cols, n_filas = leer_diseño(EXCEL_DISEÑO, catalogo)
    print(f"  Parcela: {n_cols}×{n_filas}m  |  {sum(len(c) for _,_,c in datos)} plantas\n")

    rng = random.Random(42)
    meshes = []

    # Suelo
    sv,sn,st = suelo_verts(n_cols, n_filas)
    meshes.append({"verts":sv,"norms":sn,"tris":st,"color":COLOR_SUELO})

    cnt={"CA":0,"MI":0,"ME":0,"MG":0}
    for (col, fila, cods) in datos:
        n_plantas=len(cods)
        for i,cod in enumerate(cods):
            esp=catalogo[cod]; estrato=esp["estrato"]
            altura=rng.uniform(esp["alt_min"],esp["alt_max"])
            r_copa=altura*COPA_RATIO.get(estrato,0.30)*rng.uniform(0.88,1.12)
            h_tronco=altura*rng.uniform(0.35,0.42)
            r_tronco=max(0.02, r_copa*0.09)

            if n_plantas==1:   ox,oz=rng.uniform(-.22,.22),rng.uniform(-.22,.22)
            elif n_plantas==2:
                a=math.pi*i+rng.uniform(-.2,.2); ox=math.cos(a)*.26; oz=math.sin(a)*.26
            else:
                a=2*math.pi*i/3+rng.uniform(-.15,.15); ox=math.cos(a)*.26; oz=math.sin(a)*.26

            px=col+0.5+ox; pz=fila+0.5+oz

            # Tronco
            tv,tn,tt=cilindro_verts(px,0,pz,r_tronco,h_tronco)
            meshes.append({"verts":tv,"norms":tn,"tris":tt,"color":COLOR_TRONCO_BASE})

            # Copa
            ev,en,et=esfera_verts(px,h_tronco+r_copa*0.55,pz,r_copa)
            meshes.append({"verts":ev,"norms":en,"tris":et,"color":COLOR_COPA[estrato]})

            cnt[estrato]+=1

    nombres={"CA":"Caméfitas","MI":"Microfaner.","ME":"Mesofaner.","MG":"Megafaner."}
    for k,n in cnt.items():
        print(f"    {nombres[k]:12s}: {n:4d} plantas")

    glb=build_glb(meshes)
    out=os.path.join(_DISEÑOS_DIR,"HyphaPod_3D.glb")
    with open(out,"wb") as f: f.write(glb)
    print(f"\n  ✅  {out}  ({len(glb)//1024} KB)")
    print(f"\n  → Arrastra el archivo a https://3dviewer.net")
    print(f"    (un solo archivo, sin .mtl)\n")

if __name__=="__main__":
    main()
